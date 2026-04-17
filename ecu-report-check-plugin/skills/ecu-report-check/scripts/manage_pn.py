"""
ECU 件号对照表维护脚本

用法:
  # 查看当前件号表
  python manage_pn.py --show --pn-file part_number.xlsx

  # 更新单个件号 (dry-run)
  python manage_pn.py --pn-file part_number.xlsx --update \
    --controller FLSMU --did F1A0 --value S000004325007 --dry-run

  # 从 JSON 文件批量更新
  python manage_pn.py --pn-file part_number.xlsx --input-type json --input-file update.json

  # 删除某个控制器的某个DID件号
  python manage_pn.py --pn-file part_number.xlsx --delete --controller FLSMU --did F1A0

输出 JSON 到 stdout。
"""

import os
import sys
import json
import shutil
import argparse

import openpyxl


# ─── 常量 ───

VALID_DIDS = ["F18E", "F193", "F180", "F104", "F102", "F105", "F1A0", "F103", "F17F"]
DEFAULT_CONTROLLER_COL = {
    "FLSPU": 8, "RLSPU": 9, "RRSPU": 10, "FLSMU": 11, "FRSMU": 12
}


# ─── 核心函数 ───

def load_pn_table(xlsx_path: str) -> dict:
    """加载件号表为结构化 dict。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 读取控制器名
    ctrl_cols = {}
    for col_idx in range(8, 13):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            ctrl_cols[col_idx] = str(val).strip()
        else:
            ctrl_cols[col_idx] = DEFAULT_CONTROLLER_COL.get(
                list(DEFAULT_CONTROLLER_COL.keys())[col_idx - 8], f"Col{col_idx}"
            )

    col_to_ctrl = {v: k for k, v in ctrl_cols.items()}

    # 读取件号数据
    did_rows = {}  # DID -> row number in Excel
    data = {}
    for ctrl_name in ctrl_cols.values():
        data[ctrl_name] = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        did_cell = row[2]
        if not did_cell.value:
            continue
        did = str(did_cell.value).strip()
        did_rows[did] = did_cell.row
        for col_idx, ctrl_name in ctrl_cols.items():
            pn_cell = row[col_idx - 1]
            val = pn_cell.value
            if val and str(val).strip() not in ("", "-"):
                data[ctrl_name][did] = str(val).strip()

    wb.close()
    return {
        "controllers": list(ctrl_cols.values()),
        "col_to_ctrl": col_to_ctrl,
        "data": data,
        "did_rows": did_rows,
    }


def show_pn_table(xlsx_path: str) -> dict:
    """显示当前件号表内容。"""
    table = load_pn_table(xlsx_path)
    return {
        "status": "success",
        "controllers": table["controllers"],
        "data": table["data"],
    }


def backup_file(xlsx_path: str) -> str:
    """创建备份文件，返回备份路径。"""
    bak_path = xlsx_path + ".bak"
    shutil.copy2(xlsx_path, bak_path)
    return bak_path


def apply_operations(
    xlsx_path: str,
    operations: list[dict],
    dry_run: bool = False,
) -> dict:
    """
    对件号表执行一系列操作。

    operations 中每个元素:
      {"action": "update", "controller": "FLSMU", "did": "F1A0", "value": "S000004325007"}
      {"action": "delete", "controller": "FLSMU", "did": "F1A0"}
      {"action": "add_controller", "controller": "ABCPU", "dids": {"F18E": "P000033328003", ...}}
    """
    table = load_pn_table(xlsx_path)
    changes = []
    errors = []

    for op in operations:
        action = op.get("action", "")
        ctrl = op.get("controller", "")

        if action == "update":
            did = op.get("did", "")
            value = op.get("value", "")

            if did not in VALID_DIDS:
                errors.append(f"无效 DID: {did}")
                continue
            if ctrl not in table["controllers"]:
                errors.append(f"未知控制器: {ctrl}")
                continue

            old_value = table["data"].get(ctrl, {}).get(did, "(空)")
            changes.append({
                "action": "update",
                "controller": ctrl,
                "did": did,
                "old_value": old_value,
                "new_value": value,
            })

        elif action == "delete":
            did = op.get("did", "")

            if did and did not in VALID_DIDS:
                errors.append(f"无效 DID: {did}")
                continue
            if ctrl not in table["controllers"]:
                errors.append(f"未知控制器: {ctrl}")
                continue

            if did:
                old_value = table["data"].get(ctrl, {}).get(did, "(空)")
                changes.append({
                    "action": "delete",
                    "controller": ctrl,
                    "did": did,
                    "old_value": old_value,
                })
            else:
                # 删除整个控制器的所有件号
                for d, v in table["data"].get(ctrl, {}).items():
                    changes.append({
                        "action": "delete",
                        "controller": ctrl,
                        "did": d,
                        "old_value": v,
                    })

        elif action == "add_controller":
            dids = op.get("dids", {})
            if not ctrl:
                errors.append("add_controller 需要指定 controller 名称")
                continue

            changes.append({
                "action": "add_controller",
                "controller": ctrl,
                "dids": dids,
            })

        else:
            errors.append(f"未知操作: {action}")

    if errors:
        return {"status": "error", "errors": errors, "changes_preview": changes}

    if dry_run:
        return {"status": "dry_run", "changes": changes, "message": "预览模式，未实际修改"}

    # 执行修改
    if not changes:
        return {"status": "success", "changes": [], "message": "无需修改"}

    bak_path = backup_file(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 重新读取表结构
    ctrl_cols = {}
    for col_idx in range(8, 13):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            ctrl_cols[str(val).strip()] = col_idx

    did_rows = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        did_cell = row[2]
        if did_cell.value:
            did_rows[str(did_cell.value).strip()] = did_cell.row

    next_row = ws.max_row + 1

    for change in changes:
        action = change["action"]
        ctrl = change["controller"]

        if action == "update":
            did = change["did"]
            value = change["new_value"]
            col_idx = ctrl_cols.get(ctrl)
            row_idx = did_rows.get(did)
            if col_idx and row_idx:
                ws.cell(row=row_idx, column=col_idx, value=value)

        elif action == "delete":
            did = change["did"]
            col_idx = ctrl_cols.get(ctrl)
            row_idx = did_rows.get(did)
            if col_idx and row_idx:
                ws.cell(row=row_idx, column=col_idx, value="-")

        elif action == "add_controller":
            # 添加新控制器列
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=ctrl)
            ctrl_cols[ctrl] = new_col

            # 填入件号
            dids = change.get("dids", {})
            for did, value in dids.items():
                row_idx = did_rows.get(did)
                if row_idx:
                    ws.cell(row=row_idx, column=new_col, value=value)

    wb.save(xlsx_path)
    wb.close()

    return {
        "status": "success",
        "changes": changes,
        "backup_path": bak_path,
        "message": f"已修改 {len(changes)} 项，备份已保存到 {bak_path}",
    }


def process_json_input(json_input: str) -> list[dict]:
    """将 JSON 输入转化为操作列表。"""
    data = json.loads(json_input)

    # 支持单个操作或批量操作列表
    if isinstance(data, list):
        return data

    # 单个操作对象
    action = data.get("action", "update")

    if action in ("update", "set"):
        controller = data.get("controller", "")
        dids = data.get("dids", {})
        ops = []
        for did, value in dids.items():
            ops.append({
                "action": "update",
                "controller": controller,
                "did": did,
                "value": value,
            })
        return ops

    elif action == "delete":
        controller = data.get("controller", "")
        did = data.get("did", "")
        return [{"action": "delete", "controller": controller, "did": did}]

    elif action == "add_controller":
        return [data]

    return [data]


# ─── CLI ───

def main():
    parser = argparse.ArgumentParser(description="ECU 件号对照表维护")
    parser.add_argument("--pn-file", required=True, help="件号对照表路径")
    parser.add_argument("--show", action="store_true", help="显示当前件号表")

    # 单个更新/删除
    parser.add_argument("--update", action="store_true", help="更新件号")
    parser.add_argument("--delete", action="store_true", help="删除件号")
    parser.add_argument("--controller", help="控制器名")
    parser.add_argument("--did", help="DID 代码")
    parser.add_argument("--value", help="件号值")

    # 批量输入
    parser.add_argument("--input-type", choices=["json"], help="输入类型")
    parser.add_argument("--input-file", help="输入文件路径 (用 - 表示 stdin)")

    # 安全选项
    parser.add_argument("--dry-run", action="store_true", help="仅预览，不实际修改")
    args = parser.parse_args()

    if not os.path.exists(args.pn_file):
        print(json.dumps({"error": True, "message": f"件号对照表不存在: {args.pn_file}"},
                         ensure_ascii=False, indent=2))
        sys.exit(1)

    # 显示模式
    if args.show:
        result = show_pn_table(args.pn_file)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(0)

    # 构造操作列表
    operations = []

    if args.update:
        if not args.controller or not args.did or args.value is None:
            print(json.dumps({"error": True, "message": "--update 需要 --controller, --did, --value"},
                             ensure_ascii=False, indent=2))
            sys.exit(1)
        operations.append({
            "action": "update",
            "controller": args.controller,
            "did": args.did,
            "value": args.value,
        })

    elif args.delete:
        if not args.controller:
            print(json.dumps({"error": True, "message": "--delete 需要 --controller"},
                             ensure_ascii=False, indent=2))
            sys.exit(1)
        operations.append({
            "action": "delete",
            "controller": args.controller,
            "did": args.did or "",
        })

    elif args.input_type == "json":
        if args.input_file:
            if args.input_file == "-":
                json_input = sys.stdin.read()
            else:
                with open(args.input_file, "r", encoding="utf-8") as f:
                    json_input = f.read()
        else:
            print(json.dumps({"error": True, "message": "--input-type json 需要 --input-file"},
                             ensure_ascii=False, indent=2))
            sys.exit(1)
        operations = process_json_input(json_input)

    else:
        print(json.dumps({"error": True, "message": "请指定操作: --show, --update, --delete, 或 --input-type json"},
                         ensure_ascii=False, indent=2))
        sys.exit(1)

    # 执行
    result = apply_operations(args.pn_file, operations, dry_run=args.dry_run)
    print(json.dumps(result, ensure_ascii=False, indent=2))

    if result.get("status") == "error":
        sys.exit(1)


if __name__ == "__main__":
    main()
